from __future__ import annotations

import hashlib
import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.observability import observed_task
from app.models import AsyncTask, AuditLog, BidSection, ComplianceItem, Document, ExportFile, Project, User
from app.services.storage import put_object_bytes


class ExportExcelError(Exception):
    def __init__(self, message: str, *, code: str = "EXCEL_EXPORT_FAILED") -> None:
        super().__init__(message)
        self.code = code


def _coerce_task_id(task_id: uuid.UUID | str) -> uuid.UUID:
    if isinstance(task_id, uuid.UUID):
        return task_id
    try:
        return uuid.UUID(str(task_id))
    except ValueError as exc:
        raise ExportExcelError("任务ID格式错误", code="INVALID_TASK_ID") from exc


def _filter_statement(task: AsyncTask):
    filters = task.input_json or {}
    stmt = select(ComplianceItem, Document.title, User.name).join(
        Document, Document.id == ComplianceItem.source_document_id
    ).outerjoin(User, User.id == ComplianceItem.owner_user_id)
    stmt = stmt.where(
        ComplianceItem.tenant_id == task.tenant_id,
        ComplianceItem.project_id == task.project_id,
        ComplianceItem.section_id == task.section_id,
        ComplianceItem.deleted_at.is_(None),
    )
    if filters.get("status"):
        stmt = stmt.where(ComplianceItem.status == filters["status"])
    if filters.get("risk_level"):
        stmt = stmt.where(ComplianceItem.risk_level == filters["risk_level"])
    if filters.get("owner_user_id"):
        stmt = stmt.where(ComplianceItem.owner_user_id == uuid.UUID(str(filters["owner_user_id"])))
    if filters.get("item_type"):
        stmt = stmt.where(ComplianceItem.item_type == filters["item_type"])
    return stmt.order_by(ComplianceItem.risk_level.desc(), ComplianceItem.created_at.asc())


def _status_label(value: str) -> str:
    return {
        "draft": "草稿",
        "pending_confirm": "待确认",
        "confirmed": "已确认",
        "needs_material": "缺材料",
        "rejected": "不适用",
        "superseded": "已替代",
    }.get(value, value)


def _risk_label(value: str) -> str:
    return {"high": "高", "medium": "中", "low": "低"}.get(value, value)


def _type_label(value: str) -> str:
    return {
        "qualification": "资格要求",
        "mandatory_response": "强制响应",
        "format": "格式要求",
        "deadline": "截止时间",
        "scoring": "评分办法",
        "reference_info": "参考信息",
        "technical_response": "技术响应",
        "other": "其他",
    }.get(value, value)


def _filter_label(filters: dict[str, Any] | None) -> str:
    if not filters:
        return "全部矩阵项"
    labels: list[str] = []
    if filters.get("status"):
        labels.append(f"状态={_status_label(str(filters['status']))}")
    if filters.get("risk_level"):
        labels.append(f"风险={_risk_label(str(filters['risk_level']))}")
    if filters.get("owner_user_id"):
        labels.append("责任人=已筛选")
    if filters.get("item_type"):
        labels.append(f"类型={_type_label(str(filters['item_type']))}")
    return "；".join(labels) if labels else "全部矩阵项"


def _build_workbook(
    rows: list[tuple[ComplianceItem, str | None, str | None]],
    *,
    snapshot_meta: dict[str, str | int],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合规矩阵快照"

    sheet.merge_cells("A1:J1")
    sheet["A1"] = "合规矩阵快照"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A2:J2")
    sheet["A2"] = "本文件为导出时点的只读快照，仅用于归档、会议和外部协作；最新状态以投标 Agent 平台为准，确认记录和责任分工请回到平台查看。"
    sheet["A2"].font = Font(color="9A3412")
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFF7E6")
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)

    meta_rows = [
        ("项目", snapshot_meta["project_name"], "标段", snapshot_meta["section_name"]),
        ("导出人", snapshot_meta["exported_by"], "导出时间", snapshot_meta["exported_at"]),
        ("筛选条件", snapshot_meta["filters"], "快照行数", snapshot_meta["row_count"]),
    ]
    for row_index, (left_key, left_value, right_key, right_value) in enumerate(meta_rows, start=3):
        sheet.cell(row=row_index, column=1, value=left_key).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=left_value)
        sheet.cell(row=row_index, column=5, value=right_key).font = Font(bold=True)
        sheet.cell(row=row_index, column=6, value=right_value)

    headers = [
        "序号",
        "类型",
        "风险",
        "状态",
        "强制项",
        "责任人",
        "招标要求",
        "响应建议",
        "证据来源",
        "原文证据",
    ]
    header_row = 7
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=column_index, value=header)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, (item, document_title, owner_name) in enumerate(rows, start=1):
        source = document_title or "招标文件"
        if item.source_page_no:
            source = f"{source} P{item.source_page_no}"
        row_index = header_row + index
        values = [
            index,
            _type_label(item.item_type),
            _risk_label(item.risk_level),
            _status_label(item.status),
            "是" if item.is_mandatory else "否",
            owner_name or "未指派",
            item.requirement_text,
            item.response_suggestion or "",
            source,
            item.evidence_text or "",
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    widths = [8, 14, 10, 12, 10, 14, 56, 42, 28, 56]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 36
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _add_export_audit(
    db: Session,
    task: AsyncTask,
    *,
    action: str,
    after_json: dict[str, Any],
    severity: str = "info",
) -> None:
    db.add(
        AuditLog(
            tenant_id=task.tenant_id,
            project_id=task.project_id,
            section_id=task.section_id,
            actor_user_id=task.created_by,
            actor_type="worker",
            action=action,
            object_type="async_task",
            object_id=task.id,
            after_json=after_json,
            reason="导出合规矩阵快照用于归档或外部协作",
            severity=severity,
        )
    )


@observed_task("excel_export")
def execute_compliance_matrix_excel_export_task(
    db: Session,
    task_id: uuid.UUID | str,
) -> dict[str, str | int]:
    task_uuid = _coerce_task_id(task_id)
    task = db.get(AsyncTask, task_uuid)
    if task is None or task.task_type != "excel_export":
        raise ExportExcelError("Excel 导出任务不存在", code="TASK_NOT_FOUND")

    if task.status == "succeeded" and task.output_json and task.output_json.get("export_file_id"):
        return {"status": "already_succeeded", "export_file_id": task.output_json["export_file_id"]}

    now = datetime.now(UTC)
    task.status = "running"
    task.started_at = task.started_at or now
    task.progress = 20
    db.commit()

    try:
        rows = list(db.execute(_filter_statement(task)).all())
        project = db.get(Project, task.project_id)
        section = db.get(BidSection, task.section_id) if task.section_id else None
        exported_by = db.scalar(select(User.name).where(User.id == task.created_by)) or "未知用户"
        exported_at = datetime.now(UTC)
        snapshot_meta = {
            "project_name": project.name if project else "未知项目",
            "section_name": section.name if section else "全部标段",
            "exported_by": exported_by,
            "exported_at": exported_at.strftime("%Y-%m-%d %H:%M:%S UTC"),
            "filters": _filter_label(task.input_json),
            "row_count": len(rows),
        }
        data = _build_workbook(rows, snapshot_meta=snapshot_meta)
        content_hash = hashlib.sha256(data).hexdigest()
        export_id = uuid.uuid4()
        file_name = f"合规矩阵快照-{exported_at.strftime('%Y%m%d%H%M%S')}.xlsx"
        object_key = (
            f"tenant/{task.tenant_id}/project/{task.project_id}/section/{task.section_id}/"
            f"exports/{export_id}/{file_name}"
        )
        put_object_bytes(
            bucket=settings.minio_bucket,
            object_key=object_key,
            data=data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        export_file = ExportFile(
            id=export_id,
            tenant_id=task.tenant_id,
            project_id=task.project_id,
            section_id=task.section_id,
            task_id=task.id,
            export_type="compliance_matrix_excel",
            file_name=file_name,
            bucket=settings.minio_bucket,
            object_key=object_key,
            sha256=content_hash,
            filter_json=task.input_json,
            source_snapshot_json={
                "row_count": len(rows),
                "item_ids": [str(row[0].id) for row in rows],
                "snapshot_note": "只读快照；最新状态以平台为准",
                "exported_by": exported_by,
                "exported_at": exported_at.isoformat(),
                "filters_label": snapshot_meta["filters"],
            },
            status="available",
            created_by=task.created_by,
        )
        db.add(export_file)
        task.status = "succeeded"
        task.progress = 100
        task.output_json = {
            "export_file_id": str(export_file.id),
            "file_name": file_name,
            "sha256": content_hash,
            "row_count": len(rows),
            "snapshot_note": "只读快照；最新状态以平台为准",
        }
        task.finished_at = datetime.now(UTC)
        _add_export_audit(
            db,
            task,
            action="export.excel_succeeded",
            after_json=task.output_json,
        )
        db.commit()
        return {
            "status": "succeeded",
            "export_file_id": str(export_file.id),
            "row_count": len(rows),
        }
    except Exception as exc:
        error_code = exc.code if isinstance(exc, ExportExcelError) else "EXCEL_EXPORT_FAILED"
        task.status = "failed"
        task.progress = 100
        task.error_code = error_code
        task.error_message = str(exc)
        task.finished_at = datetime.now(UTC)
        _add_export_audit(
            db,
            task,
            action="export.excel_failed",
            after_json={"error_code": error_code, "error_message": str(exc)},
            severity="warning",
        )
        db.commit()
        return {"status": "failed", "error_code": error_code}
